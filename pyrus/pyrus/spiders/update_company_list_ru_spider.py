"""下载、更新公司俄语列表"""

from datetime import date
from io import BytesIO
from zipfile import BadZipFile

import scrapy
import pymysql
from scrapy import signals
from openpyxl import load_workbook
from scrapy.spidermiddlewares.httperror import HttpError
from twisted.internet.error import DNSLookupError
from twisted.internet.error import TimeoutError, TCPTimedOutError
from twisted.internet.error import ConnectionRefusedError
from twisted.web._newclient import ResponseNeverReceived

from ..items import CompanyItem, ProfileDetailItem


class UpdateCompanyListSpider(scrapy.Spider):
    """补全公司本国名称、ipo_date等信息"""
    name = 'update_company_list_ru'

    # 应明确至此处的路径
    start_urls = [
        'https://www.moex.com/a1600'
    ]

    industry_url = 'http://fs.moex.com/files/16036'

    # 字段映射表
    mate = {
        'Дата регистрации': 'established_date',
        'Дата включения в список': 'ipo_date',
        'Эмитент': 'name_origin',
        'Торговый код': 'security_code',
        'ИНН': 'itin'
    }

    @classmethod
    def from_crawler(cls, crawler, *args, **kwargs):
        spider = super(UpdateCompanyListSpider, cls).from_crawler(
            crawler, *args, **kwargs
        )
        crawler.signals.connect(spider.spider_opened, signals.spider_opened)
        crawler.signals.connect(spider.spider_closed, signals.spider_closed)
        return spider

    def spider_opened(self, spider):
        self.logger.info('Opening spider %s...', spider.name)
        conn = pymysql.connect(**self.settings['DBARGS'])
        with conn.cursor() as cursor:
            cursor.execute("""\
                select code, security_code, name_origin, industry from \
                company where country_code_listed='RUS'\
            """)
            records = cursor.fetchall()
        conn.close()

        self.companies = {
            it['security_code']: (
                it['code'], it['industry'], it['name_origin'])
            for it in records
        }

        if records:
            NUMBER = slice(3, None)  # 公司code数字编号区
            self.max_code_num = int(max(it['code'] for it in records)[NUMBER])
        else:
            self.max_code_num = 10000

        self.total_updated = 0

    def spider_closed(self, spider):
        self.logger.info(
            'Closing spider %s..., %d updated', spider.name, self.total_updated
        )

    def parse(self, response):
        today = date.today()
        year = str(today.year)
        month = str(today.month) if today.month > 9 else '0' + str(today.month)
        doc_url = response.xpath(
            "//*[contains(text(), '{}.{}')]/@href".format(month, year)
        ).extract_first()
        if doc_url:
            yield scrapy.Request(
                doc_url,
                callback=self.parse_companies,
                errback=self.errback_scraping
            )
        else:
            self.logger.error('Failed to parse url of xlsx.')

    def parse_companies(self, response):
        try:
            wb = load_workbook(BytesIO(response.body))
            ws = wb.worksheets[0]
            headers = [cell.value for cell in ws[2]]
            for row in ws.iter_rows(min_row=2):
                values = [cell.value for cell in row]
                data = dict(zip(headers, values))
                item = CompanyItem()
                for key, val in data.items():
                    if key in self.mate:
                        try:
                            item[self.mate[key]] = val
                        except KeyError:
                            pass
                if item['security_code'] in self.companies:
                    name = self.companies[item['security_code']][2]
                    if item['name_origin'] != name:
                        self.total_updated += 1
                        item['code'] = self.companies[item['security_code']][0]
                        yield ProfileDetailItem(
                            name='itin_rus', display_label='ИНН',
                            company_code=item['code'], value=data['ИНН']
                        )
                        yield item
        except BadZipFile:
            self.logger.error('Failed to load instrument file.')

        yield scrapy.Request(
            self.industry_url,
            callback=self.parse_industries,
            errback=self.errback_scraping
        )

    def parse_industries(self, response):
        try:
            wb = load_workbook(BytesIO(response.body))
            ws = wb.worksheets[0]
            for row in ws.iter_rows(min_row=2):
                vals = [cell.value for cell in row]
                if (vals[1] in self.companies and
                        vals[3] != self.companies[vals[1]][1]):
                    yield CompanyItem(
                        code=self.companies[vals[1]][0],
                        industry=vals[3]
                    )
                    self.total_updated += 1
        except BadZipFile:
            self.logger.error('Failed to load industry file.')

    def errback_scraping(self, failure):
        req_url = failure.request.url
        if failure.check(HttpError):
            response = failure.value.response
            self.logger.error('HttpError %s on %s', response.status, req_url)
        elif failure.check(DNSLookupError):
            self.logger.error('DNSLookupError on %s', req_url)
        elif failure.check(ConnectionRefusedError):
            self.logger.error('ConnectionRefusedError on %s', req_url)
        elif failure.check(TimeoutError, TCPTimedOutError):
            self.logger.error('TimeoutError on %s', req_url)
        elif failure.check(ResponseNeverReceived):
            self.logger.error('ResponseNeverReceived on %s', req_url)
        else:
            self.logger.error('UnpectedError on %s', req_url)
            self.logger.error(repr(failure))
