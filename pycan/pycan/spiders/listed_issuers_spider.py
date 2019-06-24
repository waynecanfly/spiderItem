"""从归档（MiG Archives）文件中提取公司列表"""

from io import BytesIO
from zipfile import BadZipFile

import scrapy
import pymysql
from scrapy import signals
from openpyxl import load_workbook
from dateutil.parser import parse as parse_datetime
from scrapy.spidermiddlewares.httperror import HttpError
from twisted.internet.error import DNSLookupError
from twisted.internet.error import TimeoutError, TCPTimedOutError
from twisted.internet.error import ConnectionRefusedError
from twisted.web._newclient import ResponseNeverReceived

from ..items import CompanyItem, ProfileDetailItem


class ListedIssuersSpider(scrapy.Spider):
    name = 'listed_issuers'

    start_urls = [
        'https://www.tsx.com/listings/current-market-statistics/mig-archives'
    ]

    captions = [
        {
            'Exchange': 'exchange_market_code',
            'Name': 'name_en',
            'Root Ticker': 'security_code',
            'SP_Type': 'security_type',
            'Sector': 'sector_code',
            'Date of  TSX Listing YYYYMMDD': 'ipo_date',
            'Place of Incorporation C=Canada U=USA F=Foreign': (
                'country_code_origin'
            )
        },
        {
            'Exchange': 'exchange_market_code',
            'Name': 'name_en',
            'Root Ticker': 'security_code',
            'Sector': 'sector_code',
            'Date of Listing': 'ipo_date'
        }
    ]

    countries = {
        'C': 'CAN',
        'U': 'USA',
        'F': None
    }

    @classmethod
    def from_crawler(cls, crawler, *args, **kwargs):
        spider = super(ListedIssuersSpider, cls).from_crawler(
            crawler, *args, **kwargs
        )
        crawler.signals.connect(spider.spider_opened, signals.spider_opened)
        crawler.signals.connect(spider.spider_closed, signals.spider_closed)
        return spider

    def spider_opened(self, spider):
        self.logger.info('Opening spider %s...', spider.name)
        conn = pymysql.connect(**self.settings['DBARGS'])
        with conn.cursor() as cursor:
            cursor.execute("""\
                select code, security_code, exchange_market_code, status from \
                company where country_code_listed='CAN'\
            """)
            records = cursor.fetchall()
        conn.close()

        self.companies = {}
        for it in records:
            id_ = it['exchange_market_code'], it['security_code']
            self.companies[id_] = it['code'], it['status']

        if records:
            NUMBER = slice(3, None)  # 公司code数字编号区
            self.max_code_num = int(max(it['code'] for it in records)[NUMBER])
        else:
            self.max_code_num = 10000

        self.total_new = 0

    def spider_closed(self, spider):
        self.logger.info(
            'Closing spider %s..., %d new', spider.name, self.total_new
        )

    def parse(self, response):
        try:
            doc_href = response.xpath(
                "//a[text()='TSX/TSXV Listed Issuers']/..//a/@href"
            ).extract()[1]
            yield response.follow(
                doc_href,
                callback=self.parse_listed_issuers,
                errback=self.errback_scraping
            )
        except IndexError:
            self.logger.error("Can't find listed issuers info")

    def parse_listed_issuers(self, response):
        try:
            wb = load_workbook(BytesIO(response.body), read_only=True)
            labels_row, start_row = 7, 8
            for ws in wb.worksheets:
                labels = [
                    cell.value.replace('\n', ' ') for cell in ws[labels_row]
                    if isinstance(cell.value, str)
                ]
                names = [
                    it.replace(' ', '_').lower() + '_mig_can' for it in labels]
                for each in self.captions:
                    if set(each.keys()).issubset(set(labels)):
                        indexes = {
                            labels.index(it): each[it] for it in each
                        }
                        for row in ws.iter_rows(min_row=start_row):
                            item = CompanyItem()
                            profiles = []
                            for index, cell in enumerate(row):
                                if cell.value:
                                    try:
                                        item[indexes[index]] = cell.value
                                    except KeyError:
                                        profiles.append(
                                            ProfileDetailItem(
                                                name=names[index],
                                                display_label=labels[index],
                                                value=cell.value,
                                                data_type='string'
                                            )
                                        )
                            try:
                                item['country_code_origin'] = self.countries[
                                    item['country_code_origin']
                                ]
                            except KeyError:
                                pass
                            company = (
                                item['exchange_market_code'],
                                item['security_code']
                            )
                            if company not in self.companies:
                                self.max_code_num += 1
                                item['code'] = 'CAN' + str(self.max_code_num)
                                item['name_origin'] = item['name_en']
                                if 'ipo_date' in item:
                                    item['ipo_date'] = parse_datetime(
                                        str(item['ipo_date']))
                                self.companies[company] = (item['code'], None)
                                for p_item in profiles:
                                    p_item['company_code'] = item['code']
                                    yield p_item
                                yield item
                        break
                else:
                    self.logger.error(
                        'Failed finding captions for listed issuers')
        except BadZipFile:
            self.logger.error(
                'Listed issuers may redirect to %s', response.url)

    def errback_scraping(self, failure):
        req_url = failure.request.url
        if failure.check(HttpError):
            response = failure.value.response
            self.logger.error('HttpError %s on %s', response.status, req_url)
        elif failure.check(DNSLookupError):
            self.logger.error('DNSLookupError on %s', req_url)
        elif failure.check(ConnectionRefusedError):
            self.logger.error('ConnectionRefusedError on %s', req_url)
        elif failure.check(TimeoutError, TCPTimedOutError):
            self.logger.error('TimeoutError on %s', req_url)
        elif failure.check(ResponseNeverReceived):
            self.logger.error('ResponseNeverReceived on %s', req_url)
        else:
            self.logger.error('UnpectedError on %s', req_url)
            self.logger.error(repr(failure))
