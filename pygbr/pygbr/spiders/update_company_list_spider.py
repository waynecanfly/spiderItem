"""更新上市公司列表"""

from io import BytesIO

import scrapy
import pymysql
from scrapy import signals
from openpyxl import load_workbook
from scrapy.spidermiddlewares.httperror import HttpError
from twisted.internet.error import DNSLookupError
from twisted.internet.error import TimeoutError, TCPTimedOutError
from twisted.internet.error import ConnectionRefusedError
from twisted.web._newclient import ResponseNeverReceived

from ..items import CompanyItem, CompanyDataSourceItem


class UpdateCompanyListSpider(scrapy.Spider):
    name = 'update_company_list'

    start_url = (
        'https://www.londonstockexchange.com/statistics/companies-and-issuers/'
        'instruments-defined-by-mifir-identifiers-list-on-lse.xlsx'
    )
    search_url = (
        'http://www.londonstockexchange.com/exchange/searchengine/search.html?'
        'lang=en&q={}'
    )

    @classmethod
    def from_crawler(cls, crawler, *args, **kwargs):
        spider = super(UpdateCompanyListSpider, cls).from_crawler(
            crawler, *args, **kwargs
        )
        crawler.signals.connect(spider.spider_opened, signals.spider_opened)
        crawler.signals.connect(spider.spider_closed, signals.spider_closed)
        return spider

    def spider_opened(self, spider):
        self.logger.info('Opening spider %s...', spider.name)
        conn = pymysql.connect(**self.settings['DBARGS'])
        with conn.cursor() as cursor:
            cursor.execute("""\
                select code, name_origin, security_code from company where \
                country_code_listed='GBR'\
                """)
            records = cursor.fetchall()
        conn.close()

        self.names = [it['name_origin'] for it in records]
        self.symbols = [it['security_code'] for it in records]
        if records:
            NUMBER = slice(3, None)  # 公司code数字编号区
            self.max_code_num = int(max(it['code'] for it in records)[NUMBER])
        else:
            self.max_code_num = 10000

        self.total_new = 0

    def spider_closed(self, spider):
        self.logger.info(
            'Closing spider %s..., %d new companies', spider.name,
            self.total_new
        )

    def start_requests(self):
        yield scrapy.Request(
            self.start_url,
            callback=self.parse,
            errback=self.errback_scraping,
            meta={'download_timeout': 300}
        )

    def parse(self, response):
        if not response.headers[b'Content-Type'].decode().endswith('sheet'):
            return

        wb = load_workbook(BytesIO(response.body))
        ws = wb.worksheets[1]
        start_row, end_row = 9, ws.max_row - 1
        cols_mapping = {
            0: 'security_code',
            1: 'name_origin',
            3: 'isin',
            9: 'currency_code'
        }
        for row in ws.iter_rows(min_row=start_row, max_row=end_row):
            company = {}
            for index, cell in enumerate(row):
                if index in cols_mapping:
                    company[cols_mapping[index]] = cell.value
            if (company['name_origin'] not in self.names and
                    company['security_code'] not in self.symbols):
                self.names.append(company['name_origin'])
                self.max_code_num += 1
                company['code'] = 'GBR' + str(self.max_code_num)
                yield scrapy.Request(
                    self.search_url.format(company['security_code']),
                    callback=self.parse_company_url,
                    errback=self.errback_scraping,
                    meta={'company': company}
                )

    def parse_company_url(self, response):
        result = response.xpath(
            "//*[@class='search_results_list']/table/tbody/tr")
        security_code = response.meta['company']['security_code']
        for each in result:
            symbol = each.xpath('./td[1]/strong/text()').extract_first()
            url = each.xpath('./td[2]/a/@href').extract_first()
            if symbol == security_code:
                yield scrapy.Request(
                    url,
                    callback=self.parse_company_info,
                    errback=self.errback_scraping,
                    meta={'company': response.meta['company']}
                )
            break
        else:
            self.logger.warning('Cant\'t find info for %s', security_code)

    def parse_company_info(self, response):
        company = response.meta['company']
        website = response.xpath(
            "//*[text()='Company website']/following-sibling::td/a/@href"
        ).extract_first()
        company['website_url'] = website
        country = response.xpath(
            "//*[text()='Country of share register']/following-sibling::td/"
            "text()").extract_first()
        company['country_code_origin'] = country

        self.total_new += 1
        yield CompanyItem(company)
        yield CompanyDataSourceItem(
            company_id=company['code'], company_name=company['name_origin'],
            security_code=company['security_code'], download_link=response.url,
        )

    def errback_scraping(self, failure):
        req_url = failure.request.url
        if failure.check(HttpError):
            response = failure.value.response
            self.logger.error('HttpError %s on %s', response.status, req_url)
        elif failure.check(DNSLookupError):
            self.logger.error('DNSLookupError on %s', req_url)
        elif failure.check(ConnectionRefusedError):
            self.logger.error('ConnectionRefusedError on %s', req_url)
        elif failure.check(TimeoutError, TCPTimedOutError):
            self.logger.error('TimeoutError on %s', req_url)
        elif failure.check(ResponseNeverReceived):
            self.logger.error('ResponseNeverReceived on %s', req_url)
        else:
            self.logger.error('UnpectedError on %s', req_url)
            self.logger.error(repr(failure))
