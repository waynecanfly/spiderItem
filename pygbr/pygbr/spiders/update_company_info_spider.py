"""更新上市公司信息"""

from io import BytesIO
from collections import defaultdict
from itertools import groupby
from operator import itemgetter
from datetime import datetime

import scrapy
import pymysql
from scrapy import signals
from openpyxl import load_workbook
from scrapy.spidermiddlewares.httperror import HttpError
from twisted.internet.error import DNSLookupError
from twisted.internet.error import TimeoutError, TCPTimedOutError
from twisted.internet.error import ConnectionRefusedError
from twisted.web._newclient import ResponseNeverReceived

from ..items import ProfileDetailItem


class UpdateCompanyInfoSpider(scrapy.Spider):
    name = 'update_company_info'

    start_url = (
        'https://www.londonstockexchange.com/statistics/companies-and-issuers/'
        'instruments-defined-by-mifir-identifiers-list-on-lse.xlsx'
    )

    @classmethod
    def from_crawler(cls, crawler, *args, **kwargs):
        spider = super(UpdateCompanyInfoSpider, cls).from_crawler(
            crawler, *args, **kwargs
        )
        crawler.signals.connect(spider.spider_opened, signals.spider_opened)
        crawler.signals.connect(spider.spider_closed, signals.spider_closed)
        return spider

    def spider_opened(self, spider):
        self.logger.info('Opening spider %s...', spider.name)
        conn = pymysql.connect(**self.settings['DBARGS'])
        with conn.cursor() as cursor:
            cursor.execute("""\
                select ta.company_id, ta.security_code, ta.download_link, \
                tb.value, tc.name, tc.id as profile_id from \
                company_data_source ta left join (select * from \
                company_profile_detail where company_code like 'GBR%') tb on \
                ta.company_id = tb.company_code left join \
                company_profile_definition tc on \
                tb.company_profile_definition_id = tc.id where \
                ta.company_id like 'GBR%' order by company_id\
                """)
            records = cursor.fetchall()
        conn.close()

        self.companies = defaultdict(dict)
        for symbol, entries in groupby(records, itemgetter('security_code')):
            entries = list(entries)
            self.companies[symbol]['company_code'] = entries[0]['company_id']
            self.companies[symbol]['info_url'] = entries[0]['download_link']
            self.companies[symbol]['profiles'] = {
                it['name']: it['value'] for it in entries
            }

        self.total_updated = 0

    def spider_closed(self, spider):
        self.logger.info(
            'Closing spider %s..., %d updated', spider.name, self.total_updated
        )

    def start_requests(self):
        yield scrapy.Request(
            self.start_url,
            callback=self.parse,
            errback=self.errback_scraping,
            meta={'download_timeout': 300}
        )

    def parse(self, response):
        if not response.headers[b'Content-Type'].decode().endswith('sheet'):
            return

        updated = False
        wb = load_workbook(BytesIO(response.body))
        ws = wb.worksheets[1]
        start_row, end_row = 9, ws.max_row - 1
        labels = [cell.value for cell in ws[8]]
        names = [x.lower().replace(' ', '_') + '_gbr' for x in labels]
        for row in ws.iter_rows(min_row=start_row, max_row=end_row):
            vals = [cell.value for cell in row]
            symbol = vals[0]
            if symbol and symbol in self.companies:
                company_code = self.companies[symbol]['company_code']
                profiles = self.companies[symbol]['profiles']
                for index, name in enumerate(names):
                    if name not in profiles:  # 创建
                        updated = True
                        yield ProfileDetailItem(
                            company_code=company_code, parent_id=None,
                            company_profile_definition_id=name,
                            label=labels[index], value=vals[index],
                            gmt_create=datetime.now()
                        )
                    # handle datetime, float and None object
                    elif profiles[name] != str(vals[index]):
                        updated = True
                        yield ProfileDetailItem(
                            company_code=company_code, parent_id=None,
                            company_profile_definition_id=name,
                            value=str(vals[index]), gmt_update=datetime.now()
                        )

                yield scrapy.Request(
                    self.companies[symbol]['info_url'],
                    callback=self.parse_info,
                    errback=self.errback_scraping,
                    meta={
                        'company': company_code,
                        'profiles': profiles,
                        'updated': updated
                    }
                )

    def parse_info(self, response):
        company = response.meta['company']
        profiles = response.meta['profiles']

        updated = False
        captions = ['Company Information', 'Trading Information']
        for it in captions:
            entries = response.xpath("//*[@summary='{}']/tbody/tr".format(it))
            for each in entries:
                label = each.xpath('./td[1]/text()').extract_first().strip()
                value = each.xpath('./td[2]/text()').extract_first('').strip()
                name = label.lower().replace(' ', '_') + '_gbr'
                if name not in profiles:  # 创建
                    updated = True
                    yield ProfileDetailItem(
                        company_code=company, label=label, value=value,
                        parent_label=it,
                        parent_id=it.lower().replace(' ', '_') + '_gbr',
                        company_profile_definition_id=name,
                        gmt_create=datetime.now()
                    )
                elif value != profiles[name]:  # 更新
                    updated = True
                    yield ProfileDetailItem(
                        company_code=company, value=value,
                        parent_id=it.lower().replace(' ', '_') + '_gbr',
                        company_profile_definition_id=name,
                        gmt_update=datetime.now()
                    )

        if updated or response.meta['updated']:
            self.total_updated += 1

    def errback_scraping(self, failure):
        req_url = failure.request.url
        if failure.check(HttpError):
            response = failure.value.response
            if response.status != 404:
                self.logger.error(
                    'HttpError %s on %s', response.status, req_url)
        elif failure.check(DNSLookupError):
            self.logger.error('DNSLookupError on %s', req_url)
        elif failure.check(ConnectionRefusedError):
            self.logger.error('ConnectionRefusedError on %s', req_url)
        elif failure.check(TimeoutError, TCPTimedOutError):
            self.logger.error('TimeoutError on %s', req_url)
        elif failure.check(ResponseNeverReceived):
            self.logger.error('ResponseNeverReceived on %s', req_url)
        else:
            self.logger.error('UnpectedError on %s', req_url)
            self.logger.error(repr(failure))
